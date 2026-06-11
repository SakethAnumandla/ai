#!/usr/bin/env python3
"""Build API_200OK_LIST.xlsx from curl test report or route inventory."""
from __future__ import annotations

import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "API_200OK_LIST.xlsx"
REPORT = ROOT / "docs" / "API_CURL_TEST_REPORT.md"
BASE = os.environ.get("EXPENSE_API_BASE", "http://127.0.0.1:8000").rstrip("/")

ROW_RE = re.compile(
    r"^\|\s*\d+\s*\|\s*(GET|POST|PUT|PATCH|DELETE)\s*\|\s*`([^`]+)`\s*\|\s*(\d+|SKIP)\s*\|"
)


def parse_curl_report(path: Path) -> tuple[list[dict], list[dict]]:
    """Return (ok_rows, failed_rows) from API_CURL_TEST_REPORT.md."""
    if not path.is_file():
        return [], []
    ok_rows: list[dict] = []
    failed_rows: list[dict] = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        m = ROW_RE.match(line)
        if not m:
            continue
        method, route_path, status = m.groups()
        code = int(status) if status.isdigit() else 0
        row = {
            "name": f"{method} {route_path}",
            "method": method,
            "path": route_path,
            "full_url": f"{BASE}{route_path}",
            "status_code": status,
            "status_text": "OK" if 200 <= code < 300 else status,
        }
        if 200 <= code < 300:
            ok_rows.append(row)
        else:
            failed_rows.append(row)
    return ok_rows, failed_rows


def all_routes() -> list[dict]:
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))
    from app.main import app

    rows = []
    for route in app.routes:
        if not hasattr(route, "methods") or not hasattr(route, "path"):
            continue
        if route.path in {"/docs", "/openapi.json", "/redoc", "/docs/oauth2-redirect"}:
            continue
        for method in sorted(route.methods - {"HEAD", "OPTIONS"}):
            rows.append(
                {
                    "name": f"{method} {route.path}",
                    "method": method,
                    "path": route.path,
                    "full_url": f"{BASE}{route.path}",
                    "status_code": "",
                    "status_text": "Registered route",
                }
            )
    rows.sort(key=lambda r: (r["path"], r["method"]))
    return rows


def write_excel(ok_rows: list[dict], all_rows: list[dict], failed_rows: list[dict] | None = None) -> None:
    wb = Workbook()
    headers = ["S.No", "API Name", "Method", "Path", "Full URL", "Status Code", "Status Text"]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)

    ws = wb.active
    ws.title = "200 OK APIs"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(ok_rows, 1):
        ws.append([i, r["name"], r["method"], r["path"], r["full_url"], r["status_code"], r["status_text"]])

    ws2 = wb.create_sheet("All Registered APIs")
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(1, c)
        cell.fill = hfill
        cell.font = hfont
    for i, r in enumerate(all_rows, 1):
        ws2.append([i, r["name"], r["method"], r["path"], r["full_url"], r["status_code"], r["status_text"]])

    meta = wb.create_sheet("Summary")
    meta.append(["Field", "Value"])
    meta.append(["Base URL", BASE])
    meta.append(["Test Date", datetime.now(timezone.utc).strftime("%Y-%m-%d")])
    meta.append(["Source", str(REPORT.name)])
    meta.append(["200 OK (2xx) APIs", len(ok_rows)])
    meta.append(["Failed / Skipped", len(failed_rows or [])])
    meta.append(["Total Registered APIs", len(all_rows)])
    meta.append(["Generated At (UTC)", datetime.now(timezone.utc).isoformat()])

    if failed_rows:
        ws3 = wb.create_sheet("Failures")
        ws3.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws3.cell(1, c)
            cell.fill = hfill
            cell.font = hfont
        for i, r in enumerate(failed_rows, 1):
            ws3.append([i, r["name"], r["method"], r["path"], r["full_url"], r["status_code"], r["status_text"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            w = min(max(len(str(c.value or "")) for c in col) + 2, 70)
            sheet.column_dimensions[col[0].column_letter].width = w

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Saved {len(ok_rows)} rows -> {OUTPUT}")


def main() -> int:
    ok, failed = parse_curl_report(REPORT)
    if not ok:
        print(f"No 2xx rows in {REPORT} — run scripts/run_every_api_curl.py first", file=sys.stderr)
        return 1
    all_rows = all_routes()
    write_excel(ok, all_rows, failed)
    print(f"2xx: {len(ok)} | non-2xx: {len(failed)} | registered: {len(all_rows)}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
