#!/usr/bin/env python3
"""Export all APIs returning 2xx to Excel (local seeded smoke test or production Newman)."""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("Install openpyxl: python3.11 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
COLLECTION = ROOT / "postman" / "Bizwy_Expense_API.postman_collection.json"
OUTPUT = ROOT / "docs" / "API_200OK_LIST.xlsx"

ENV_VARS = {
    "base_url": "https://api.bizwy.in",
    "session_id": "api-test-session01",
    "expense_draft_id": "1",
    "expense_submitted_id": "2",
    "expense_empty_draft_id": "3",
    "expense_rejected_id": "4",
    "expense_thumb_id": "5",
    "expense_file_id": "1",
    "expense_approval_id": "1",
    "policy_id": "1",
    "policy_deletable_id": "2",
    "claim_id": "1",
    "claim_approval_id": "1",
    "ocr_bill_id": "1",
    "ocr_batch_id": "1",
    "job_id": "1",
    "finance_report_job_id": "2",
    "snapshot_a_id": "1",
    "snapshot_b_id": "2",
    "alert_id": "1",
    "bulk_export_id": "00000000-0000-0000-0000-000000000001",
    "review_token": "api-test-review-token-0001",
    "category": "meals_entertainment",
    "financial_year": "FY2025-26",
}


def run_local_smoke() -> list[dict]:
    """Run full route inventory against seeded in-memory DB (TestClient)."""
    os.environ.setdefault("DATABASE_URL", "sqlite://")
    os.environ.setdefault("PADDLE_OCR_PRELOAD", "0")
    os.environ.setdefault("REDIS_ENABLED", "0")
    os.environ.setdefault("TESTING", "1")
    os.environ.setdefault("OCR_TEST_BYPASS", "1")
    os.environ.setdefault("OPENAI_API_KEY", "")

    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))

    from fastapi.testclient import TestClient
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool

    import app.ai.models as _ai_models  # noqa: F401
    import app.finance.models as _finance_models  # noqa: F401
    import app.database as db_module
    from app.database import Base, get_db
    from app.main import app
    from app.models import AIChatSession  # noqa: F401
    from tests import api_test_context
    from tests.seed_data import reset_and_seed
    from tests.test_all_api_routes import (
        ALL_API_ROUTES,
        SUCCESS_STATUS,
        _call_route,
    )

    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    db_module.engine = engine
    db_module.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

    db = db_module.SessionLocal()
    try:
        api_test_context.CURRENT_SEED = reset_and_seed(db)
        db.commit()

        def override_get_db():
            try:
                yield db
            finally:
                pass

        app.dependency_overrides[get_db] = override_get_db
        rows: list[dict] = []
        with TestClient(app) as client:
            for method, path in ALL_API_ROUTES:
                try:
                    resp = _call_route(client, method, path)
                    code = resp.status_code
                    rows.append(
                        {
                            "name": f"{method} {path}",
                            "method": method,
                            "path": path,
                            "full_url": f"http://127.0.0.1:8000{path}",
                            "status_code": code,
                            "status_text": str(code),
                            "response_time_ms": None,
                            "success_2xx": code in SUCCESS_STATUS,
                        }
                    )
                except Exception as exc:
                    rows.append(
                        {
                            "name": f"{method} {path}",
                            "method": method,
                            "path": path,
                            "full_url": f"http://127.0.0.1:8000{path}",
                            "status_code": 0,
                            "status_text": str(exc)[:80],
                            "response_time_ms": None,
                            "success_2xx": False,
                        }
                    )
        app.dependency_overrides.clear()
        return rows
    finally:
        db.close()


def run_newman_production(base_url: str) -> list[dict]:
    if not COLLECTION.is_file():
        raise FileNotFoundError(
            f"Collection missing: {COLLECTION}. Run: python3.11 scripts/generate_postman_collection.py"
        )
    env = dict(ENV_VARS)
    env["base_url"] = base_url

    with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as out:
        out_path = out.name

    cmd = [
        "npx", "-y", "newman", "run", str(COLLECTION),
        "--timeout-request", "120000",
        "--delay-request", "150",
        "--reporters", "json",
        "--reporter-json-export", out_path,
    ]
    for k, v in env.items():
        cmd.extend(["--env-var", f"{k}={v}"])

    print(f"Running Newman against {base_url} (may take several minutes)...")
    subprocess.run(cmd, cwd=str(ROOT), check=False)
    data = json.loads(Path(out_path).read_text(encoding="utf-8"))
    Path(out_path).unlink(missing_ok=True)

    rows: list[dict] = []
    for ex in (data.get("run") or {}).get("executions") or []:
        item = ex.get("item") or {}
        request = ex.get("request") or {}
        method = (request.get("method") or "").upper()
        url_obj = request.get("url") or {}
        raw = url_obj.get("raw", "") if isinstance(url_obj, dict) else str(url_obj)
        path = raw.replace(base_url, "").split("?")[0] or "/"
        response = ex.get("response") or {}
        code = response.get("code")
        if code is None and ex.get("requestError"):
            code = 0
        rows.append(
            {
                "name": item.get("name") or f"{method} {path}",
                "method": method,
                "path": path,
                "full_url": raw,
                "status_code": code,
                "status_text": response.get("status") or "",
                "response_time_ms": response.get("responseTime"),
                "success_2xx": code is not None and 200 <= int(code) < 300,
            }
        )
    return rows


def write_excel(rows: list[dict], output: Path, *, source: str, base_url: str) -> None:
    ok_rows = [r for r in rows if r["success_2xx"]]
    fail_rows = [r for r in rows if not r["success_2xx"]]

    wb = Workbook()
    headers = [
        "S.No",
        "API Name",
        "Method",
        "Path",
        "Full URL",
        "Status Code",
        "Status Text",
        "Response Time (ms)",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, ncol: int) -> None:
        for col in range(1, ncol + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws) -> None:
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    ws_ok = wb.active
    ws_ok.title = "200 OK APIs"
    ws_ok.append(headers)
    style_header(ws_ok, len(headers))
    for i, r in enumerate(ok_rows, 1):
        ws_ok.append(
            [i, r["name"], r["method"], r["path"], r["full_url"],
             r["status_code"], r["status_text"], r["response_time_ms"]]
        )
    autosize(ws_ok)

    ws_all = wb.create_sheet("All APIs")
    ws_all.append(headers)
    style_header(ws_all, len(headers))
    for i, r in enumerate(rows, 1):
        ws_all.append(
            [i, r["name"], r["method"], r["path"], r["full_url"],
             r["status_code"], r["status_text"], r["response_time_ms"]]
        )
    autosize(ws_all)

    ws_fail = wb.create_sheet("Failed APIs")
    fail_headers = headers + ["Note"]
    ws_fail.append(fail_headers)
    for col in range(1, len(fail_headers) + 1):
        cell = ws_fail.cell(row=1, column=col)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = Font(color="FFFFFF", bold=True)
    for i, r in enumerate(fail_rows, 1):
        ws_fail.append(
            [i, r["name"], r["method"], r["path"], r["full_url"],
             r["status_code"], r["status_text"], r["response_time_ms"],
             "Did not return 2xx"]
        )
    autosize(ws_fail)

    meta = wb.create_sheet("Summary")
    meta.append(["Field", "Value"])
    meta.append(["Test Source", source])
    meta.append(["Base URL", base_url])
    meta.append(["Generated At (UTC)", datetime.now(timezone.utc).isoformat()])
    meta.append(["Total APIs Tested", len(rows)])
    meta.append(["200 OK (2xx) Count", len(ok_rows)])
    meta.append(["Failed Count", len(fail_rows)])
    meta.append(["Output File", str(output.name)])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Wrote {len(ok_rows)} / {len(rows)} APIs (2xx) to {output}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export 2xx APIs to Excel")
    parser.add_argument(
        "--production",
        action="store_true",
        help="Test live production via Newman (slow; needs postman collection)",
    )
    parser.add_argument(
        "--base-url",
        default="https://api.bizwy.in",
        help="Production base URL when using --production",
    )
    parser.add_argument(
        "-o", "--output",
        default=str(OUTPUT),
        help="Output .xlsx path",
    )
    args = parser.parse_args()
    output = Path(args.output)

    if args.production:
        rows = run_newman_production(args.base_url)
        source = f"Production Newman ({args.base_url})"
        base = args.base_url
    else:
        print("Running local seeded API smoke test...")
        rows = run_local_smoke()
        source = "Local TestClient (seeded SQLite)"
        base = "http://127.0.0.1:8000"

    if not rows:
        print("No API results collected.", file=sys.stderr)
        return 1

    write_excel(rows, output, source=source, base_url=base)
    return 0


if __name__ == "__main__":
    sys.exit(main())
